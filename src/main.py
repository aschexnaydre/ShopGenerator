#!/usr/bin/env python3
# -*-coding: utf-8-*-
__author__ = "Matthew Dickson"
__credits__ = ["Jacob Valdez of RoleplayingTips.com"]

__license__ = "MIT"
__version__ = "1.5"

import openpyxl
import tkinter
import random
import math
from tkinter import filedialog, Button, Entry, Checkbutton, Radiobutton, Label, IntVar, StringVar, messagebox
import os.path
import errno


def main(sett_size, shop_size, count, data_path, return_path):

    if sett_size is 0:
        tkinter.messagebox.showerror("Argument Error", "Must selected a settlement size!")
        return
    elif shop_size is 0:
        tkinter.messagebox.showerror("Argument Error", "Must selected a shop size!")
        return
    elif data_path is None or not data_path.endswith('.xlsx'):
        tkinter.messagebox.showerror("File Error", "Must select a .xlsx file!")
        return

    xls_flag = flag.get()
    ret_table = []
    gen_count = 1
    if count is not "" and math.floor(float(count)) is not 0:
        gen_count = math.floor(float(count))

    my_book = openpyxl.load_workbook(filename=data_path, read_only=True)
    main_table = my_book['Sheet1']

    city_max = get_city_max(sett_size)
    shop_weight = get_shop_weight(shop_size)
    city_id = sett_size

    if xls_flag is 1:
        ret_book = Workbook()

    for i in range(0,gen_count):
        for row in main_table.rows:
            if row[0].value is None or row[0].value == 'Item':  # Cell is null or is a header
                continue
            item_weight = get_item_weight(row[1].value)
            base_price = row[2].value
            selling_price = get_list_price(city_id, base_price)
            num_avail = get_avail_count(item_weight, shop_weight, city_max, selling_price)
            if num_avail > 0:
                ret_table.append([row[0].value, str(num_avail), str(selling_price)])

        ret_table = handle_currency_overflow(city_max, ret_table, shop_size)
        ret_table.sort()
        handle_print(ret_table)
        if xls_flag is 1:
            handle_xls(ret_table, ret_book, i, return_path, sett_size, shop_size)
        ret_table = []


def get_shop_string(shop_size):
    switcher = {
        1: "Tiny",
        2: "Small",
        3: "Medium",
        4: "Large"
    }
    return switcher.get(shop_size, "")


def get_city_string(sett_size):
    switcher = {
        1: "Thorp",
        2: "Hamlet",
        3: "Village",
        4: "SmallTown",
        5: "LargeTown",
        6: "SmallCity",
        7: "LargeCity",
        8: "Metropolis"
    }
    return switcher.get(sett_size, "")


def get_city_max(city_type):
    switcher = {
        1: 4000,
        2: 10000,
        3: 20000,
        4: 80000,
        5: 30000,
        6: 1500000,
        7: 4000000,
        8: 10000000,
    }
    return switcher.get(city_type, "")


def get_item_weight(item_weight):
    switcher = {
        "C": 2,
        "U": 5,
        "R": 7,
        "E": 9,
    }
    return switcher.get(item_weight)


def get_shop_weight(shop_size):
    switcher = {
        1: 4,
        2: 6,
        3: 8,
        4: 10,
    }
    return switcher.get(shop_size, "")


def handle_xls(my_table, my_book, i, return_path, sett_size, shop_size):
    my_sheet = my_book.create_sheet("Shop " + str(i + 1))
    my_sheet.append(["Item", "Quantity", "Price (cp)"])
    for row in my_table:
        my_sheet.append([row[0], int(row[1]), int(row[2])])
    try:
        os.makedirs(return_path)
    except OSError as err:
        if errno.errorcode[err.errno] == 'EEXIST':  # If the error is because the file exists
            pass
        else:
            tkinter.messagebox.showerror("File Error", err)
    else:
        tkinter.messagebox.showerror("File Error", "Something went wrong!")
    city_string = get_city_string(sett_size)
    shop_string = get_shop_string(shop_size)
    if not os.path.isfile(return_path + city_string + "_" + shop_string + "Shop.xlsx"):
        my_book.save(return_path + city_string + "_" + shop_string + "Shop.xlsx")
    else:
        book_id = 1
        while os.path.isfile(return_path + city_string + "_" + shop_string + "Shop" + "_" + str(book_id) + ".xlsx"):
            book_id += 1
        my_book.save(return_path + city_string + "_" + shop_string + "Shop" + "_" + str(book_id) + ".xlsx")


#  Helps keep the total item count reasonable (~50-100) in thorps, hamlets, and villages. Also helps small settlements
#  have varying shop sizes
def handle_currency_overflow(city_max, avail_items, shop_size):
    mod_factor = math.floor(math.log10(city_max) * (shop_size - 1) * 10**(math.floor(math.log10(city_max)) - 1))
    city_max += shop_size * shop_size * mod_factor  # Accounts for larger shops having more of one item
    items_total_cost = 0
    rand_avail_items = list(avail_items)
    random.shuffle(rand_avail_items)
    my_return = []
    for item in rand_avail_items:
        if int(item[2]) + items_total_cost <= city_max:
            items_total_cost += int(item[2])
            my_return.append(item)
    return my_return


def handle_print(ret_table):
    template = "{0:50}{1:10}{2:11}"
    print(template.format("Item", "Quantity", "Price (cp)"))
    print(template.format("----", "--------", "----------"))
    for r in ret_table:
        print(((template.format(*r)).encode('utf-8')).decode('utf-8'))
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")


def get_avail_count(item_weight, shop_weight, city_max, list_price):
    if list_price <= city_max:
        return randint(0, shop_weight) - item_weight
    else:
        return 0


def get_list_price(city_id, base_price):
    return math.ceil((abs(randint(6,21)- math.ceil(city_id/randint(1,3)))+1)/10*int(base_price))


def handle_files(data_path, return_path, root):
    file = filedialog.askopenfilename()
    data_path.set(file)
    dir = os.path.dirname(file)
    ret_dir = os.path.abspath(os.path.join(dir, 'shops'))
    return_path.set(ret_dir + "\\")
    length = len(data_path.get())
    if length > 22:
        Label(root, text="..." + data_path.get()[length-19:], width=22).grid(row=9, column=1)
    else:
        Label(root, text=data_path.get(), width=22).grid(row=9, column=1)
    

root = tkinter.Tk()
root.title("Shop Maker v" + str(__version__))
root.geometry('275x250')
flag = IntVar()
sett_size = IntVar()
shop_size = IntVar()
data_path = StringVar()
return_path = StringVar()

Label(root, text='Settlement Size:').grid(row=0, column=0)
Radiobutton(text='Thorp', variable=sett_size, value=1).grid(row=2, column=0)
Radiobutton(text='Hamlet', variable=sett_size, value=2).grid(row=3, column=0)
Radiobutton(text='Village', variable=sett_size, value=3).grid(row=4, column=0)
Radiobutton(text='Small Town', variable=sett_size, value=4).grid(row=5, column=0)
Radiobutton(text='Large Town', variable=sett_size, value=5).grid(row=6, column=0)
Radiobutton(text='Small City', variable=sett_size, value=6).grid(row=7, column=0)
Radiobutton(text='Large City', variable=sett_size, value=7).grid(row=8, column=0)
Radiobutton(text='Metropolis', variable=sett_size, value=8).grid(row=9, column=0)

Label(root, text='Shop Size:').grid(row=0, column=1)
Radiobutton(text='Tiny', variable=shop_size, value=1).grid(row=2, column=1)
Radiobutton(text='Small', variable=shop_size, value=2).grid(row=3, column=1)
Radiobutton(text='Medium', variable=shop_size, value=3).grid(row=4, column=1)
Radiobutton(text='Large', variable=shop_size, value=4).grid(row=5, column=1)

Label(root, text='Number to Make:').grid(row=6, column=1)
count = Entry(root, width=7)
count.grid(row=7, column=1)

Checkbutton(root, text='Save to .xlsx file?', variable=flag).grid(row=10, column=0)
Button(root, text='Data File...', command=lambda: handle_files(data_path, return_path, root)).grid(row=8,column=1)

Button(root, text='Go', command=lambda:
        main(sett_size.get(), shop_size.get(), count.get(), data_path.get(), return_path.get())).grid(row=10, column=1)


root.mainloop()
